"""openpyxl extraction adapter."""
from __future__ import annotations

from openpyxl import load_workbook  # type: ignore


def extract(file_path: str) -> str | None:
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value) for value in row]
                if any(cells):
                    lines.append(f"[{worksheet.title}]\t" + "\t".join(cells))
        workbook.close()
        return "\n".join(lines)
    except Exception:
        return None
