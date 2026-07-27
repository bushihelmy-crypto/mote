"""openpyxl extraction adapter."""

from __future__ import annotations

from openpyxl import load_workbook  # type: ignore

from mote.runtime.fileops.document_budgets import BoundedTextSink


def extract(file_path: str, *, sink: BoundedTextSink) -> None:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        separator = ""
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value) for value in row]
                if any(cells):
                    sink.write(separator)
                    sink.write(f"[{worksheet.title}]\t")
                    sink.write("\t".join(cells))
                    separator = "\n"
    finally:
        workbook.close()
